#!/usr/bin/env python3
"""Phase 1 — Import der wöchentlichen Sender-Hitliste (xlsx) in kuenstler.db.

Findet die Headerzeile dynamisch, normalisiert Künstlernamen und pflegt die
Datenbank idempotent (Upsert). Priorität wird nie abgesenkt.

Namenskonventionen (Sender-Export):
- Nur der erste Künstler steht im Komma-Format: „Giesinger, Max"
- „Nachname, The" → „The Nachname" (Band)
- feat. / x / vs. → getrennte Künstler
- „&": ohne Komma-Segment = ein Act (Hall & Oates);
  mit Komma-Segment und Artikel-Suffix = ein Act (Lewis, Huey & The News);
  sonst getrennte Künstler

Nutzung:
    python3 import_hitliste.py sender_hitliste.xlsx [--prio-spalte "Prio"]
"""

import argparse
import re
import sqlite3
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from config import DB_PATH

# ---------------------------------------------------------------------------
# Datenbank
# ---------------------------------------------------------------------------

SCHEMA = """
CREATE TABLE IF NOT EXISTS kuenstler (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT    NOT NULL UNIQUE COLLATE NOCASE,
    prioritaet      TEXT    DEFAULT 'C',   -- A / B / C
    typ             TEXT,                  -- Person / Band / NULL (Phase 2)
    geburtsdatum    TEXT,
    todesdatum      TEXT,
    gruendungsjahr  INTEGER,
    wikidata_id     TEXT,
    musicbrainz_id  TEXT,
    quellen         TEXT,                  -- JSON
    verifiziert     INTEGER DEFAULT 0,
    first_seen      TEXT    NOT NULL,
    last_seen       TEXT    NOT NULL
);

CREATE TABLE IF NOT EXISTS titel (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    kuenstler_id    INTEGER NOT NULL REFERENCES kuenstler(id),
    titel           TEXT    NOT NULL,
    voe_monat       INTEGER,
    voe_jahr        INTEGER,
    first_seen      TEXT    NOT NULL,
    last_seen       TEXT    NOT NULL,
    UNIQUE(kuenstler_id, titel)
);

CREATE TABLE IF NOT EXISTS mitglieder (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    band_id         INTEGER NOT NULL REFERENCES kuenstler(id),
    person_name     TEXT    NOT NULL,
    rolle           TEXT,
    kuratiert       INTEGER DEFAULT 0,
    UNIQUE(band_id, person_name)
);

CREATE TABLE IF NOT EXISTS news (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    kuenstler_id    INTEGER NOT NULL REFERENCES kuenstler(id),
    typ             TEXT    NOT NULL,       -- album, single, tour, tv, news
    ereignis_datum  TEXT,
    text            TEXT    NOT NULL,
    quellen         TEXT,                   -- JSON
    verifiziert     INTEGER DEFAULT 0,
    gefunden_am     TEXT    NOT NULL,
    hash            TEXT    NOT NULL UNIQUE
);
"""


def init_db(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.executescript(SCHEMA)
    conn.commit()
    return conn


# ---------------------------------------------------------------------------
# Header-Erkennung
# ---------------------------------------------------------------------------

KNOWN_HEADERS = {"interpret", "künstler", "artist", "titel", "title", "song"}


def find_header_row(ws):
    """Sucht in den ersten 20 Zeilen nach einer Zeile mit bekannten Headern."""
    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=False), 1):
        values = [str(c.value).strip().lower() for c in row if c.value is not None]
        if any(v in KNOWN_HEADERS for v in values):
            col_map = {}
            for c in row:
                if c.value is not None:
                    col_map[str(c.value).strip().lower()] = c.column - 1
            return row_idx, col_map
    return None, None


# ---------------------------------------------------------------------------
# Namens-Normalisierung
# ---------------------------------------------------------------------------

FEAT_PATTERN = re.compile(
    r"\s+(?:feat\.?|ft\.?|featuring)\s+", re.IGNORECASE
)
VS_PATTERN = re.compile(r"\s+(?:vs\.?|versus)\s+", re.IGNORECASE)
X_PATTERN = re.compile(r"\s+x\s+", re.IGNORECASE)

ARTIKEL = {"the", "die", "les", "los", "las", "de", "el", "la"}


def _flip_comma_name(name: str) -> str:
    """'Nachname, Vorname' → 'Vorname Nachname'; 'Nachname, The' → 'The Nachname'."""
    if "," not in name:
        return name.strip()
    parts = [p.strip() for p in name.split(",", 1)]
    if len(parts) == 2 and parts[1]:
        if parts[1].lower() in ARTIKEL:
            return f"{parts[1]} {parts[0]}"
        return f"{parts[1]} {parts[0]}"
    return name.strip()


def _is_article_suffix(rest: str) -> bool:
    """Prüft ob der &-Rest mit einem Artikel beginnt (→ ein Act)."""
    first_word = rest.strip().split()[0].lower() if rest.strip() else ""
    return first_word in ARTIKEL


def parse_artists(raw: str) -> list[str]:
    """Zerlegt den Interpreten-String in einzelne Künstlernamen."""
    if not raw or not raw.strip():
        return []

    parts = FEAT_PATTERN.split(raw)
    result = []
    for part in parts:
        for sub in VS_PATTERN.split(part):
            for xsub in X_PATTERN.split(sub):
                result.append(xsub.strip())

    final = []
    for segment in result:
        if "&" not in segment:
            final.append(_flip_comma_name(segment))
            continue

        if "," in segment.split("&")[0]:
            comma_part, amp_rest = segment.split("&", 1)
            if _is_article_suffix(amp_rest):
                flipped = _flip_comma_name(comma_part.strip())
                final.append(f"{flipped} & {amp_rest.strip()}")
            else:
                final.append(_flip_comma_name(comma_part.strip()))
                for a in amp_rest.split("&"):
                    a = a.strip()
                    if a:
                        final.append(_flip_comma_name(a))
        else:
            final.append(segment.strip())

    return [n for n in final if n]


# ---------------------------------------------------------------------------
# Priorität
# ---------------------------------------------------------------------------

PRIO_RANK = {"A": 3, "B": 2, "C": 1}


def higher_prio(existing: str, new: str) -> str:
    """Gibt die höhere Priorität zurück (nie absenken)."""
    return existing if PRIO_RANK.get(existing, 0) >= PRIO_RANK.get(new, 0) else new


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def import_hitliste(xlsx_path: Path, prio_spalte: str | None = None):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    header_row, col_map = find_header_row(ws)
    if header_row is None:
        print("FEHLER: Keine Header-Zeile gefunden.", file=sys.stderr)
        sys.exit(1)

    interpret_col = None
    for key in ("interpret", "künstler", "artist"):
        if key in col_map:
            interpret_col = col_map[key]
            break
    if interpret_col is None:
        print("FEHLER: Keine Interpret-Spalte gefunden.", file=sys.stderr)
        sys.exit(1)

    titel_col = None
    for key in ("titel", "title", "song"):
        if key in col_map:
            titel_col = col_map[key]
            break

    prio_col = None
    if prio_spalte and prio_spalte.lower() in col_map:
        prio_col = col_map[prio_spalte.lower()]

    voe_col = None
    for key in ("vö", "voe", "veröffentlichung", "release", "jahr"):
        if key in col_map:
            voe_col = col_map[key]
            break

    conn = init_db(DB_PATH)
    heute = date.today().isoformat()
    stats = {"titel": 0, "kuenstler_neu": 0, "kuenstler_gesehen": 0}

    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    wb.close()

    for row in rows:
        if not row or len(row) <= interpret_col:
            continue
        raw_interpret = row[interpret_col]
        if not raw_interpret:
            continue

        raw_titel = row[titel_col] if titel_col is not None and len(row) > titel_col else None
        raw_prio = str(row[prio_col]).strip().upper() if prio_col is not None and len(row) > prio_col and row[prio_col] else "C"
        if raw_prio not in PRIO_RANK:
            raw_prio = "C"

        voe_monat, voe_jahr = None, None
        if voe_col is not None and len(row) > voe_col and row[voe_col]:
            voe_raw = str(row[voe_col]).strip()
            m = re.match(r"(\d{1,2})[./](\d{4})", voe_raw)
            if m:
                voe_monat, voe_jahr = int(m.group(1)), int(m.group(2))
            else:
                m = re.match(r"(\d{4})", voe_raw)
                if m:
                    voe_jahr = int(m.group(1))

        artists = parse_artists(str(raw_interpret))

        for artist_name in artists:
            existing = conn.execute(
                "SELECT id, prioritaet FROM kuenstler WHERE name = ? COLLATE NOCASE",
                (artist_name,),
            ).fetchone()

            if existing:
                kid, old_prio = existing
                new_prio = higher_prio(old_prio, raw_prio)
                conn.execute(
                    "UPDATE kuenstler SET last_seen = ?, prioritaet = ? WHERE id = ?",
                    (heute, new_prio, kid),
                )
                stats["kuenstler_gesehen"] += 1
            else:
                cur = conn.execute(
                    "INSERT INTO kuenstler (name, prioritaet, first_seen, last_seen) VALUES (?, ?, ?, ?)",
                    (artist_name, raw_prio, heute, heute),
                )
                kid = cur.lastrowid
                stats["kuenstler_neu"] += 1

            if raw_titel:
                titel_str = str(raw_titel).strip()
                conn.execute(
                    """INSERT INTO titel (kuenstler_id, titel, voe_monat, voe_jahr, first_seen, last_seen)
                       VALUES (?, ?, ?, ?, ?, ?)
                       ON CONFLICT(kuenstler_id, titel) DO UPDATE
                       SET last_seen = excluded.last_seen,
                           voe_monat = COALESCE(excluded.voe_monat, voe_monat),
                           voe_jahr  = COALESCE(excluded.voe_jahr, voe_jahr)""",
                    (kid, titel_str, voe_monat, voe_jahr, heute, heute),
                )
                stats["titel"] += 1

    conn.commit()
    conn.close()

    total_k = stats["kuenstler_neu"] + stats["kuenstler_gesehen"]
    print(f"Import abgeschlossen: {stats['titel']} Titel, "
          f"{total_k} Künstler ({stats['kuenstler_neu']} neu), "
          f"DB: {DB_PATH}")


def main():
    parser = argparse.ArgumentParser(description="Sender-Hitliste importieren")
    parser.add_argument("xlsx", type=Path, help="Pfad zur xlsx-Datei")
    parser.add_argument("--prio-spalte", default=None, help="Name der Prioritäts-Spalte")
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"FEHLER: Datei nicht gefunden: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    import_hitliste(args.xlsx, args.prio_spalte)


if __name__ == "__main__":
    main()
